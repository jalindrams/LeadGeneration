"""
Micraft Growth Engine - Import Call Feedback from Excel
Reads a call-status spreadsheet from the sales team, matches rows to DB leads
by normalized phone number, and records the outcome:

  - lead.call_status / lead.response_status updated
  - lead_feedback row created per sheet entry (Module 2: Sales Feedback Loop)
  - contact timestamps set

Status mapping (raw rep notes -> canonical):
  wrong_contact   invalid / wrong number / not valid / no incoming / out of
                  service / switched off / cannot be connected ...
  not_interested  not interested / no need / not required / already have ...
  no_response     didn't pick / busy / cut / voicemail / no call ...
  interested      interested / call back / follow up ...

Usage:
  python scripts/import_call_feedback.py "C:\\path\\to\\leads.xlsx" [--contact-date 2026-07-16]
"""

import argparse
import os
import re
import sys
from datetime import datetime
from collections import Counter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

from app.database import SessionLocal
from app.models import Lead, LeadFeedback
from app.processing.dedup import normalize_phone

# Keyword rules, checked in order — first match wins
STATUS_RULES = [
    ("not_interested", ["not interested", "not intrested", "no need", "not required",
                        "no requiremnt", "no requirement", "already have", "listen and cut",
                        "no interest", "dont want", "don't want"]),
    ("interested", ["interested", "call back", "callback", "follow up", "follow-up",
                    "send details", "whatsapp details", "asked for"]),
    ("wrong_contact", ["invalid", "invaild", "wrong", "not valid", "no incoming",
                       "not in service", "out of service", "switch", "cannot be connected",
                       "cant be reached", "can't be reached", "not reachable", "no service",
                       "does not exist", "doesnt exist"]),
    ("no_response", ["didnt pick", "didn't pick", "did not pick", "not picked", "not picking",
                     "dint pick", "didnt picked", "busy", "cut", "voicemail", "didnt received",
                     "no call", "no answer", "ringing", "language barrier", "no response"]),
]

# When multiple reps called the same lead, keep the most informative outcome
PRECEDENCE = {"converted": 5, "interested": 4, "not_interested": 3, "wrong_contact": 2, "no_response": 1}

# response_status -> call_status for the calling UI
CALL_STATUS_MAP = {
    "interested": "interested",
    "not_interested": "not_interested",
    "wrong_contact": "closed",
    "no_response": "no_answer",
}


def map_status(raw: str) -> str | None:
    if not raw:
        return None
    text = str(raw).strip().lower()
    if not text:
        return None
    for canonical, keywords in STATUS_RULES:
        if any(k in text for k in keywords):
            return canonical
    # Bare "no" means not interested in rep shorthand
    if text in ("no", "nope"):
        return "not_interested"
    return "no_response"  # unrecognized note — least assuming bucket


def main():
    parser = argparse.ArgumentParser(description="Import call feedback from Excel")
    parser.add_argument("xlsx_path")
    parser.add_argument("--contact-date", default="2026-07-16",
                        help="Nominal date the calls were made (YYYY-MM-DD)")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    contact_dt = datetime.strptime(args.contact_date, "%Y-%m-%d")
    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)

    db = SessionLocal()
    try:
        # Build phone -> lead index for real (non-synthetic) leads
        phone_index = {}
        for lead in db.query(Lead).filter(Lead.status != "synthetic").all():
            p = normalize_phone(lead.phone or "")
            if len(p) == 10:
                phone_index.setdefault(p, lead)

        matched, unmatched, feedback_rows = 0, 0, 0
        outcome_counts = Counter()
        callers = Counter()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c).strip().lower() if c else "" for c in rows[0]]

            def col_exact(*names):
                for i, h in enumerate(header):
                    if h in names:
                        return i
                return None

            phone_i = next((i for i, h in enumerate(header) if "phone" in h), None)
            status_is = [i for i, h in enumerate(header) if "status" in h]
            status_i = status_is[-1] if status_is else None
            caller_i = col_exact("call holder", "caller", "name", "person")

            if phone_i is None or status_i is None:
                continue

            for r in rows[1:]:
                phone_raw = r[phone_i] if phone_i < len(r) else None
                status_raw = r[status_i] if status_i < len(r) else None
                if not phone_raw or not status_raw:
                    continue
                canonical = map_status(status_raw)
                if not canonical:
                    continue

                phone = normalize_phone(re.sub(r"\D", "", str(phone_raw)))
                lead = phone_index.get(phone)
                if not lead:
                    unmatched += 1
                    continue

                matched += 1
                outcome_counts[canonical] += 1
                caller = str(r[caller_i]).strip() if caller_i is not None and caller_i < len(r) and r[caller_i] else "Sales Rep"
                callers[caller.title()] += 1

                if args.dry_run:
                    continue

                # Feedback row (Module 2)
                db.add(LeadFeedback(
                    lead_id=lead.id,
                    status=canonical,
                    notes=f"[import {sheet_name}] {caller}: {str(status_raw).strip()}",
                    created_at=contact_dt,
                ))
                feedback_rows += 1

                # Lead-level rollup with precedence
                current = lead.response_status
                if not current or PRECEDENCE.get(canonical, 0) > PRECEDENCE.get(current, 0):
                    lead.response_status = canonical
                    lead.call_status = CALL_STATUS_MAP.get(canonical, "new")
                lead.feedback_count = (lead.feedback_count or 0) + 1
                lead.call_count = (lead.call_count or 0) + 1
                if not lead.first_contacted_at:
                    lead.first_contacted_at = contact_dt
                lead.last_contacted_at = contact_dt
                lead.outreach_status = "contacted"
                lead.last_outreach_channel = "call"

        if not args.dry_run:
            db.commit()

        print(f"{'DRY RUN — ' if args.dry_run else ''}Feedback import complete")
        print(f"  Rows matched to real leads : {matched}")
        print(f"  Rows unmatched (fake/missing): {unmatched}")
        print(f"  Feedback rows created      : {feedback_rows}")
        print("\nOutcomes recorded:")
        for k, v in outcome_counts.most_common():
            print(f"  {k:15s}: {v}")
        print("\nCalls by rep:")
        for k, v in callers.most_common(10):
            safe = k.encode("ascii", "replace").decode()
            print(f"  {safe:15s}: {v}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
